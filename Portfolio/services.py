import re
from datetime import datetime
from openpyxl import load_workbook
from django.db import transaction, connection
from .models import Stock, HistoricalPrice

DATE_RE = re.compile(r"^\d{4}_\d{2}_\d{2}$")

def parse_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    if s == "" or s.lower() in {"nan", "none"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None

def load_market_full(file_path, batch_size=5000, only_last_n_sheets=None):
    """
    Loads ALL rows from ALL date sheets into Stock + HistoricalPrice using batching.
    Uses ignore_conflicts to avoid slow update_or_create.
    """
    # SQLite lock reduction (safe to keep even after moving to Postgres)
    with connection.cursor() as c:
        try:
            c.execute("PRAGMA journal_mode=WAL;")
            c.execute("PRAGMA synchronous=NORMAL;")
        except Exception:
            pass

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = [s for s in wb.sheetnames if DATE_RE.match(s)]
    sheet_names.sort()  # oldest -> newest; optional

    if only_last_n_sheets:
        sheet_names = sheet_names[-only_last_n_sheets:]

    total_prices_inserted = 0
    total_sheets = 0

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col_index = {str(h).strip(): i for i, h in enumerate(header) if h}

        # required columns
        i_symbol = col_index.get("Symbol")
        i_close = col_index.get("Close")
        if i_symbol is None or i_close is None:
            continue  # skip sheet if format is unexpected

        # optional columns
        i_open = col_index.get("Open")
        i_high = col_index.get("High")
        i_low = col_index.get("Low")
        i_vol = col_index.get("Vol")
        i_turn = col_index.get("Turnover") or col_index.get("Amount")

        sheet_date = datetime.strptime(sheet_name, "%Y_%m_%d").date()
        total_sheets += 1

        symbols = set()
        rows_cache = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            symbol = row[i_symbol]
            if not symbol:
                continue
            symbol = str(symbol).strip().upper()
            symbols.add(symbol)
            open_p  = parse_float(row[i_open]) if i_open is not None else None
            high_p  = parse_float(row[i_high]) if i_high is not None else None
            low_p   = parse_float(row[i_low])  if i_low  is not None else None
            close_p = parse_float(row[i_close])

            vol = parse_float(row[i_vol]) if i_vol is not None else None
            if vol is not None and vol < 0:
                vol = 0

            turnover = parse_float(row[i_turn]) if i_turn is not None else None
            if turnover is not None and turnover < 0:
                turnover = 0

            rows_cache.append((
                 symbol,
                open_p,
                high_p,
                low_p,
                close_p,
                vol,
                turnover,
            ))
           

        if not rows_cache:
            continue

        # 2) Bulk create missing stocks
        existing = set(
            Stock.objects.filter(ticker__in=symbols).values_list("ticker", flat=True)
        )
        missing = [Stock(ticker=s) for s in symbols if s not in existing]
        if missing:
            Stock.objects.bulk_create(missing, ignore_conflicts=True)

        stock_map = dict(
            Stock.objects.filter(ticker__in=symbols).values_list("ticker", "id")
        )

        # 3) Build HistoricalPrice objects and bulk insert in batches
        buffer = []
        for symbol, op, hp, lp, cp, vol, turnover in rows_cache:
            if cp is None:
                continue

            buffer.append(HistoricalPrice(
                stock_id=stock_map[symbol],
                date=sheet_date,
                open_price=op,
                high_price=hp,
                low_price=lp,
                close_price=cp,
                volume=vol,
                turnover=turnover,
            ))

            if len(buffer) >= batch_size:
                with transaction.atomic():
                    HistoricalPrice.objects.bulk_create(
                        buffer,
                        batch_size=batch_size,
                        ignore_conflicts=True
                    )
                total_prices_inserted += len(buffer)
                buffer.clear()

        if buffer:
            with transaction.atomic():
                HistoricalPrice.objects.bulk_create(
                    buffer,
                    batch_size=batch_size,
                    ignore_conflicts=True
                )
            total_prices_inserted += len(buffer)

    wb.close()

    return {
        "message": "Full import completed",
        "sheets_processed": total_sheets,
        "prices_attempted_insert": total_prices_inserted,
        "note": "If you re-import, duplicates are ignored due to unique(stock,date) + ignore_conflicts.",
    }